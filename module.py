from bs4 import BeautifulSoup
import urllib.request
import requests
import fake_useragent
import os
import pathlib
import string
import random
import json
import re

import openpyxl





class DataPage:
    url = ''
    properties = []
    resault = {}
    resault['is'] = None
    resault['src'] = None
    resault['category'] = None
    resault['name'] = None
    resault['price'] = 0
    resault['art_namber'] = None
    resault['main'] = None
    resault['more'] = []
    resault['descrip'] = ''
    resault['json'] = None
    resault['video'] = []
    resault['docs'] = []
    resault['properties'] = []
    ua = fake_useragent.UserAgent()

    def __init__(self, url):
        self.url = url
        self.headers = {"User-Agent": self.ua.random}
        self.resault['video'] = []
        self.resault['docs'] = []




    def randomStr(self, nrandchars):
        alpha = string.ascii_letters + string.digits
        chars = ''.join(random.choice(alpha) for _ in range(nrandchars))
        return chars

    def get_files(self, link, folder='images'):
        response = requests.get(link, stream=True)
        ext = pathlib.Path(link).suffix
        if not os.path.isdir("upload"):
            os.mkdir("upload")
        if not os.path.isdir("upload/" + folder):
            os.mkdir("upload/" + folder)
        file_name = 'upload/' + folder + '/' + folder + '_' + self.randomStr(10) + ext
        file = open(file_name, 'bw')
        for chunk in response.iter_content(4096):
            file.write(chunk)
        return file_name

    def get(self):
        req = urllib.request.Request(self.url, data=None, headers=self.headers)
        req = urllib.request.urlopen(req)
        html = req.read()
        soup = BeautifulSoup(html, 'html.parser')
        propsBody = []
        try:
            propsBody = soup.find('div', {'id' : 'content_features'}).findAll('div', class_='ty-product-feature')
        except:
            pass
        if len(propsBody) > 0:
            #self.properties.clear()
            propsItems = []
            for props in propsBody:
                item = {
                    'label' : props.find('div', class_='ty-product-feature__label').get_text(strip=True),
                    'value' : props.find('div', class_='ty-product-feature__value').get_text(strip=True)
                }

                propsItems.append(item)

            self.properties = propsItems

        self.resault['json'] = ''
        if len(self.properties) > 0:
            forJson = {}
            key = 1
            for item in self.properties:
                forJson[key] = item
                key += 1

            self.resault['json']  = json.dumps(forJson, ensure_ascii=False)

        self.resault['src'] = soup.find('link', {'rel':'canonical'}).get('href')
        self.resault['name'] = soup.find('h1', {'class':'ty-product-block-title'}).get_text(strip=True)
        breadcrumbs = soup.find_all('a', class_='ty-breadcrumbs__a')
        if len(breadcrumbs)>0: self.resault['category'] = breadcrumbs[-1].getText()
        if self.resault['name'] != None:
            self.resault['is'] = True
        try:
            price = soup.find('span', class_='ty-price').span.get_text(strip=True)
            self.resault['price'] = int(price.replace(' ', ''))
        except:
            pass
        self.resault['art_namber'] = soup.find('span', {'class':'ty-control-group__item'}).get_text(strip=True)







        try:
            self.resault['descrip'] = str(soup.select('#content_description')[0].div.decode_contents())
        except:
            pass
        imagesAll = soup.find('div', class_='ty-product-img').findAll('a', class_='cm-image-previewer')
        imagesArr = []
        imagesMorre = []
        for imgItem in imagesAll:
            imagesArr.append(imgItem.get('href'))

        if len(imagesArr) > 0:
            c = 1
            for img in imagesArr:
                if c == 1: self.resault['main'] = self.get_files(img, 'main')
                if c > 1: imagesMorre.append(self.get_files(img, 'more'))
                c+=1
            if len(imagesMorre) > 0: self.resault['more'] = imagesMorre

        try:
            content_videos = soup.find('div', {'id': 'content_videos'})
            if len(content_videos)>0:

                videos = content_videos.find_all('a')
                for one in videos:
                    self.resault['video'].append(one.get_text(strip=True))
                    self.resault['video'].append(one.get('href'))
        except:
            pass
        try:
            content_attachments = soup.find('div', {'class': 'attachments'})
            if len(content_attachments)>0:

                docs = content_attachments.find_all('p', class_='attachment__item')
                for doc in docs:
                    self.resault['docs'].append(doc.get_text(strip=True))
                    self.resault['docs'].append(doc.find('a', class_='attachment__a').get('href'))
        except:
            pass

        self.resault['properties'] = self.properties

        return  self.resault


class DataGetLinks:
    url = ''
    paginLink = []
    links = []

    def __init__(self, url):
        self.url = url

    def getSoup(self, src):
        req = urllib.request.urlopen(src)
        html = req.read()
        return BeautifulSoup(html, 'html.parser')

    def getLinks(self, src):
        links = []
        soup = self.getSoup(src)
        try:
            items = soup.find_all('a', class_="product-title")
            for item in items:
                links.append(item.get('href'))
        except:
            pass
        return links

    def get(self):
        try:
            soup = self.getSoup(self.url)
            linkArr = []
            paginLink = []
            pagination = soup.find('div', class_="ty-pagination__items")
            linkArr = pagination.find_all('a', {'class': 'ty-pagination__item'})
            self.paginLink.append(self.url)
            if(len(linkArr)>0):
                for item in linkArr:
                    self.paginLink.append(item.get('href'))
        except:
            pass
        if len(self.paginLink) > 0:
            for l in self.paginLink:
                self.links.extend(self.getLinks(l))
        else:
            self.links.extend(self.getLinks(self.url))
        return (self.links)




class DataPaginations:
    url = ''
    links = []
    def __init__(self, url):
        self.url = url
    def get(self):
        req = urllib.request.urlopen(self.url)
        html = req.read()
        soup = BeautifulSoup(html, 'html.parser')
        try:
            pagination = soup.find('div', class_="ty-pagination__items")
            linkArr = pagination.find_all('a', {'class': 'ty-pagination__item'})
            self.links.append(self.url)
            for item in linkArr:
                self.links.append(item.get('href'))
        except:
            pass
        return (self.links)


class DataLinksOnPage:
    url = ''
    links = []
    def __init__(self, url):
        self.url = url

    def get(self):
        req = urllib.request.urlopen(self.url)
        html = req.read()
        soup = BeautifulSoup(html, 'html.parser')
        try:
            items = soup.find_all('a', class_="product-title")
            for item in items:
                self.links.append(item.get('href'))
        except:
            pass

        return (self.links)


class DataLinksOnAllPages:
    urls = []
    links = []
    def __init__(self, urls):
        self.urls = urls
    def get(self):
        try:
            for url in self.urls:
                req = urllib.request.urlopen(url)
                html = req.read()
                soup = BeautifulSoup(html, 'html.parser')
                items = soup.find_all('a', class_="product-title")
                for item in items:
                    self.links.append(item.get('href'))
        except:
            pass
        return (self.links)


class ExcelFile:
    data = {}
    labels = set()

    def __init__(self, data):
        self.data = data

    def get(self):
        for item in self.data:
            if len(item['properties']) > 0:
                for prop in item['properties']:
                    self.labels.add(prop['label'])

        book = openpyxl.Workbook()
        sheet = book.active

        sheet['A1'].value = 'src'
        sheet['B1'].value = 'name'
        sheet['C1'].value = 'category'
        sheet['D1'].value = 'art_namber'
        sheet['E1'].value = 'price'
        sheet['F1'].value = 'main'
        sheet['G1'].value = 'more'
        sheet['H1'].value = 'descrip'
        sheet['I1'].value = 'JSON_properties'
        sheet['J1'].value = 'video'
        sheet['K1'].value = 'docs'

        if len(self.labels) > 0:
            nam = 12
            for nameProp in self.labels:
                sheet.cell(row=1, column=nam, value=nameProp)
                nam += 1

        row = 2

        for res in self.data:
            sheet[row][0].value = res['src']
            sheet[row][1].value = res['name']
            sheet[row][2].value = res['category']
            sheet[row][3].value = res['art_namber']
            sheet[row][4].value = res['price']
            sheet[row][5].value = res['main']
            sheet[row][6].value = ','.join(res['more'])
            sheet[row][7].value = res['descrip']
            sheet[row][8].value = res['json']
            sheet[row][9].value = ','.join(res['video'])
            sheet[row][10].value = ','.join(res['docs'])
            if len(self.labels) > 0:
                namberCell = 12
                while namberCell < len(self.labels) + 12:
                    propLabel = sheet.cell(row=1, column=namberCell).value
                    for property in res['properties']:
                        if property['label'] == propLabel:
                            sheet.cell(row=row, column=namberCell).value = property['value']
                        else:
                            sheet.cell(row=row, column=namberCell)
                    namberCell += 1
            row += 1

        file_name = detNameExcelFile()
        book.save(file_name)
        book.close()

        return file_name
















# получаем орегинальное имя Excel - файла
def detNameExcelFile():
    max_namber = 1
    if os.path.isdir("upload"):
        nambers = []
        filenames = next(os.walk('upload'), (None, None, []))[2]  # [] if no file
        if len(filenames) > 0:
            for nam in filenames:
                if nam[0].isdigit(): nambers.append(int(nam[0]));
            max_namber = max(nambers) + 1
    else:
        os.mkdir("upload")
    return 'upload/'+str(max_namber)+'_resault.xlsx'


# проверяем ссылку на соответствие домену
def isLinkCorrect(url):
    domain = r'https://vezuviy.su/|https://everest-pech.com/|https://etna-pech.ru/'
    if re.match(domain, url, re.IGNORECASE):
        return True
    else:
        return False

# проверяем ссылку на соответствие URL
def isLink(url):
    url_pattern = "^https?:\\/\\/(?:www\\.)?[-a-zA-Z0-9@:%._\\+~#=]{1,256}\\.[a-zA-Z0-9()]{1,6}\\b(?:[-a-zA-Z0-9()@:%_\\+.~#?&\\/=]*)$"
    if re.match(url_pattern, url):
        return True
    else:
        return False
